"""변환 결과 → 엑셀.

셀 서식이 아니라 **셀 타입**을 제대로 넣는 것이 핵심이다. 계정코드를 텍스트로 쓰지
않으면 엑셀이 열면서 선행 0 을 지워버려, 파이썬 쪽에서 아무리 보존해도 소용이 없다.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..convert import ConversionResult
from ..fields import FieldType

MONEY_FORMAT = "#,##0;[Red](#,##0)"
"""회계 관례의 음수 표기: 빨간 괄호."""

DATE_FORMAT = "yyyy-mm-dd"
TEXT_FORMAT = "@"

_HEADER_FILL = PatternFill("solid", fgColor="E8EDEB")
_ISSUE_FILL = PatternFill("solid", fgColor="FBE9E6")


def write_xlsx(
    result: ConversionResult,
    path: str | Path,
    *,
    sheet_name: str = "변환결과",
    mark_issues: bool = True,
    issue_sheet: bool = True,
) -> Path:
    """엑셀로 내보낸다.

    ``issue_sheet=True`` 면 이슈 목록을 별도 시트로 함께 넣는다. 변환 산출물만
    받아본 사람도 무엇을 검수해야 하는지 파일 안에서 알 수 있어야 한다.
    """
    path = Path(path)
    specs = result.profile.fields

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    for col, spec in enumerate(specs, start=1):
        cell = ws.cell(row=1, column=col, value=spec.key)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    row_no = 1
    for converted in result.rows:
        if not converted.included:
            continue
        row_no += 1
        for col, spec in enumerate(specs, start=1):
            result_cell = converted.cells[spec.key]
            cell = ws.cell(row=row_no, column=col)
            cell.value = _excel_value(result_cell.value, spec.type)
            _apply_format(cell, spec.type)
            if mark_issues and result_cell.issues:
                cell.fill = _ISSUE_FILL

    ws.freeze_panes = "A2"
    for col, spec in enumerate(specs, start=1):
        ws.column_dimensions[get_column_letter(col)].width = _width(spec.key, spec.type)

    if issue_sheet and result.issues:
        _write_issue_sheet(wb, result)

    wb.save(path)
    return path


def _excel_value(value: Any, ftype: FieldType) -> Any:
    if value is None:
        return None
    if ftype is FieldType.CODE:
        # 반드시 문자열로. 셀 서식만으로는 선행 0 이 지켜지지 않는다.
        return str(value)
    if ftype is FieldType.DATE and isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return value
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    return value


def _apply_format(cell, ftype: FieldType) -> None:
    if ftype is FieldType.CODE:
        cell.number_format = TEXT_FORMAT
        cell.alignment = Alignment(horizontal="left")
    elif ftype in (FieldType.MONEY, FieldType.INTEGER, FieldType.DECIMAL):
        cell.number_format = MONEY_FORMAT if ftype is FieldType.MONEY else "#,##0.####"
        cell.alignment = Alignment(horizontal="right")
    elif ftype is FieldType.DATE:
        cell.number_format = DATE_FORMAT


def _width(key: str, ftype: FieldType) -> int:
    base = max(10, min(28, len(key) + 4))
    if ftype in (FieldType.MONEY, FieldType.DECIMAL):
        return max(base, 14)
    if ftype is FieldType.DATE:
        return max(base, 12)
    return base


def _write_issue_sheet(wb: Workbook, result: ConversionResult) -> None:
    ws = wb.create_sheet("검수필요")
    for col, title in enumerate(["원본행", "필드", "심각도", "코드", "내용"], start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL

    for i, (source_row, key, issue) in enumerate(result.issues, start=2):
        ws.cell(row=i, column=1, value=source_row)
        ws.cell(row=i, column=2, value=key)
        ws.cell(row=i, column=3, value=issue.severity.value)
        ws.cell(row=i, column=4, value=issue.code)
        ws.cell(row=i, column=5, value=issue.message)

    ws.freeze_panes = "A2"
    for col, width in enumerate((9, 16, 10, 22, 70), start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
