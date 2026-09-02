"""변환·프로파일·라이터·미리보기 통합."""

from __future__ import annotations

import csv
import json

import openpyxl

from everyfile import Profile, convert_file, infer_profile, load
from everyfile.fields import FieldType
from everyfile.profile import suggest_key
from everyfile.writers.csv_writer import _escape


def test_infers_a_usable_profile(ledger_xlsx):
    """스키마를 손으로 쓰게 하면 아무도 쓰지 않는다 — 초안이 바로 쓸만해야 한다."""
    job = load(ledger_xlsx)
    types = {f.key: f.type for f in job.profile.fields}

    assert types["entryDate"] is FieldType.DATE
    assert types["accountCode"] is FieldType.CODE
    assert types["accountName"] is FieldType.TEXT
    assert types["debit"] is FieldType.MONEY
    assert types["counterparty"] is FieldType.TEXT


def test_maps_korean_headers_to_standard_keys():
    assert suggest_key("전표일자") == "entryDate"
    assert suggest_key("계정 코드") == "accountCode"
    assert suggest_key("차변") == "debit"
    assert suggest_key("거래처명") == "counterparty"
    assert suggest_key("특이사항") == "특이사항"  # 사전에 없으면 원문 유지


def test_conversion_drops_subtotals_but_keeps_the_anchor(ledger_xlsx):
    """제외된 행도 자리를 남겨야 화면에 gap 을 그릴 수 있다."""
    job = load(ledger_xlsx)
    dropped = [r for r in job.result.rows if not r.included]

    assert len(dropped) == 1
    assert dropped[0].source_index == 12
    assert dropped[0].drop_reason == "소계·합계 행 — 데이터에서 제외"
    assert len(job.result.records) == 11


def test_conversion_applies_the_rules_end_to_end(ledger_xlsx):
    job = load(ledger_xlsx)
    records = job.result.records

    assert records[0]["entryDate"] == "2026-01-05"
    assert records[0]["accountCode"] == "0110"  # 선행 0 보존
    assert records[0]["counterparty"] is None  # "-" → null
    assert records[6]["debit"] == -240000  # (240,000) → 음수


def test_json_output_roundtrips(ledger_xlsx, tmp_path):
    target = tmp_path / "out.json"
    convert_file(ledger_xlsx, target)

    records = json.loads(target.read_text(encoding="utf-8"))
    assert len(records) == 11
    assert records[0]["accountCode"] == "0110"
    assert isinstance(records[0]["accountCode"], str)
    assert records[3]["debit"] == 34500000


def test_xlsx_output_keeps_code_as_text(ledger_xlsx, tmp_path):
    """엑셀은 셀 타입이 숫자면 서식과 무관하게 선행 0 을 지운다."""
    target = tmp_path / "out.xlsx"
    convert_file(ledger_xlsx, target)

    wb = openpyxl.load_workbook(target)
    ws = wb["변환결과"]
    header = [c.value for c in ws[1]]
    code_col = header.index("accountCode") + 1
    cell = ws.cell(row=2, column=code_col)

    assert cell.value == "0110"
    assert isinstance(cell.value, str)
    assert cell.number_format == "@"


def test_xlsx_output_carries_an_issue_sheet(ledger_xlsx, tmp_path):
    target = tmp_path / "out.xlsx"
    profile = load(ledger_xlsx).profile
    profile.field_by_key("accountCode").type = FieldType.INTEGER  # 이슈를 유발

    convert_file(ledger_xlsx, target, profile=profile)
    wb = openpyxl.load_workbook(target)
    assert "검수필요" in wb.sheetnames
    assert wb["검수필요"].max_row > 1


def test_csv_output_uses_bom_for_excel(ledger_xlsx, tmp_path):
    """BOM 이 없으면 한국어 윈도우 엑셀이 UTF-8 CSV 를 CP949 로 읽어 한글이 깨진다."""
    target = tmp_path / "out.csv"
    convert_file(ledger_xlsx, target)
    assert target.read_bytes().startswith(b"\xef\xbb\xbf")

    rows = list(csv.reader(target.read_text(encoding="utf-8-sig").splitlines()))
    assert rows[0][0] == "entryDate"
    assert rows[1][2] == "현금"


def test_csv_escapes_formula_injection():
    """외부에서 받은 파일을 변환해 배포하는 일이 잦아 이 방어가 필수다."""
    assert _escape("=cmd|'/c calc'!A1") == "'=cmd|'/c calc'!A1"
    assert _escape("@SUM(1:2)") == "'@SUM(1:2)"
    assert _escape("+1+1") == "'+1+1"
    assert _escape("-1240000") == "-1240000"  # 음수 금액은 통과
    assert _escape("현금") == "현금"


def test_profile_roundtrips_through_json(ledger_xlsx, tmp_path):
    """프로파일 저장·재사용이 제품의 핵심 가치다."""
    original = load(ledger_xlsx).profile
    path = tmp_path / "profile.json"
    original.save(path)

    restored = Profile.load(path)
    assert [f.key for f in restored.fields] == [f.key for f in original.fields]
    assert [f.type for f in restored.fields] == [f.type for f in original.fields]


def test_profile_rejects_a_newer_schema_version():
    """버저닝이 없으면 저장된 프로파일이 6개월 뒤 조용히 깨진다."""
    import pytest

    with pytest.raises(ValueError, match="최신"):
        Profile.from_dict({"name": "x", "schemaVersion": 999, "fields": []})


def test_profile_exports_json_schema(ledger_xlsx):
    schema = load(ledger_xlsx).profile.to_json_schema()
    props = schema["items"]["properties"]

    assert schema["type"] == "array"
    assert props["accountCode"] == {"anyOf": [{"type": "string"}, {"type": "null"}]}
    assert props["entryDate"]["anyOf"][0]["format"] == "date"


def test_unmapped_fields_are_reported(ledger_xlsx):
    """원본에 없는 열을 조용히 비워두면 나중에 원인을 못 찾는다."""
    profile = Profile(name="t")
    profile.fields = [
        *infer_profile(load(ledger_xlsx).table).fields,
    ]
    profile.fields.append(
        type(profile.fields[0])(key="taxCode", source="세목코드", type=FieldType.CODE)
    )

    job = load(ledger_xlsx, profile=profile)
    assert job.result.unmapped == ["taxCode"]
    assert any("taxCode" in n for n in job.result.notes)


def test_preview_pairs_source_and_output_by_anchor(ledger_xlsx):
    payload = load(ledger_xlsx).preview()

    src_rows = [r["sourceRow"] for r in payload["source"]["rows"]]
    out_rows = [r["sourceRow"] for r in payload["output"]["rows"]]
    assert src_rows == out_rows
    assert src_rows == sorted(src_rows)

    gap = next(r for r in payload["output"]["rows"] if not r["included"])
    assert gap["sourceRow"] == 12
    assert "소계" in gap["dropReason"]


def test_preview_carries_rules_and_diffs(ledger_xlsx):
    payload = load(ledger_xlsx).preview()
    row = next(r for r in payload["output"]["rows"] if r["sourceRow"] == 10)
    debit = next(c for c in row["cells"] if c["key"] == "debit")

    assert debit["value"] == -240000
    assert debit["diff"] == "changed"
    assert debit["rules"] == ["괄호→음수", "strip:thousands", "cast:number"]


def test_preview_describes_its_sampling(ledger_xlsx):
    """샘플만 보고 전체를 판단하게 하려면 무엇을 봤는지 밝혀야 한다."""
    payload = load(ledger_xlsx).preview()
    sampling = payload["sampling"]

    assert sampling["total"] == 11
    assert "전체 11행" in sampling["label"]
    assert sampling["seed"] == 20260101


def test_preview_is_deterministic(ledger_xlsx):
    """같은 파일의 미리보기가 매번 달라지면 비교가 불가능하다."""
    first = load(ledger_xlsx).preview()
    second = load(ledger_xlsx).preview()
    assert first == second


def test_dropped_rows_do_not_report_issues(ledger_xlsx):
    """소계 행의 '소계' 를 날짜로 못 읽은 건 검수 대상이 아니라 제외한 이유다."""
    job = load(ledger_xlsx)
    assert job.result.issues == []
    assert job.result.error_count == 0


def test_dropped_rows_report_issues_once_they_are_included(ledger_xlsx):
    """제외를 끄면 그 행도 출력 대상이므로 이슈가 다시 올라와야 한다."""
    profile = load(ledger_xlsx).profile
    profile.drop_subtotals = False

    job = load(ledger_xlsx, profile=profile)
    assert len(job.result.records) == 12
    assert any(row == 12 for row, _, _ in job.result.issues)
